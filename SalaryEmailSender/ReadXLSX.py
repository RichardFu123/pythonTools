"""
Shawn 11/21/2018
excel读取部分
"""

import openpyxl
import OneEmailInfo


class ReadXLSX:
    work_book = None
    salary_sheet = None
    address_sheet = None
    sheets = []

    max_row = 0
    max_col = 0
    name_col = 0
    rows_in_salary = None
    cols_in_salary = None
    row_lines_in_salary = []
    col_lines_in_salary = []

    def __init__(self, file_name):
        self.work_book = openpyxl.load_workbook(file_name)
        self.sheets = self.work_book.get_sheet_names()

        try:
            if "邮件地址" in self.sheets and "工资数据" in self.sheets:
                self.address_sheet = self.work_book.get_sheet_by_name("邮件地址")
                self.salary_sheet = self.work_book.get_sheet_by_name("工资数据")
            else:
                self.address_sheet = self.work_book.get_sheet_by_name(self.sheets[1])
                self.salary_sheet = self.work_book.get_sheet_by_name(self.sheets[0])
        except Exception:
            print("Can not get sheets.")

        try:
            self.rows_in_salary = self.salary_sheet.rows
            for row in self.rows_in_salary:
                self.row_lines_in_salary.append(col.value for col in row)

            self.cols_in_salary = self.salary_sheet.columns
            for col in self.cols_in_salary:
                self.col_lines_in_salary.append(row.value for row in col)
        except Exception:
            print("Can not read the salary table.")

        for i in range(4, len(self.row_lines_in_salary) + 1):
            if self.salary_sheet.cell(row=i, column=2).value == "合计":
                self.max_row = i - 1

        for i in range(1, len(self.col_lines_in_salary) + 1):
            if self.salary_sheet.cell(row=4, column=i).value == "实发\n工资":
                self.max_col = i
            if self.salary_sheet.cell(row=4, column=i).value == "姓名":
                self.name_col = i

    def get_email_address(self, name):
        flag = False
        for i in self.address_sheet.iter_rows():
            for j in i:
                if flag:
                    return j.value
                if j.value == name:
                    flag = True

    def get_emails_pack(self):
        pack = []

        for row in self.salary_sheet.iter_rows(min_row=5, max_row=self.max_row, max_col=self.max_col):
            person = []
            for info in row:
                person.append(str(info.value))
            name = person[self.name_col-1]
            email_address = self.get_email_address(name)

            date_str = str(self.salary_sheet.cell(row=2, column=15).value)
            title = str(self.salary_sheet.cell(row=1, column=1).value)
            manager = "HR"

            one_email = OneEmailInfo.OneEmailInfo(name)
            one_email.email_address = email_address
            one_email.date_str = date_str
            one_email.manager = manager
            one_email.title = title
            one_email.personal_part = person

            pack.append(one_email)

        return pack


if __name__ == "__main__":
    test = ReadXLSX("test.xlsx")
    packs = test.get_emails_pack()
    for one in packs:
        print(one)
